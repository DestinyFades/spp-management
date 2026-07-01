from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from typing import Dict

class ExportService:
    @staticmethod
    async def generate_excel(tree_data: Dict, total: float) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Распределение СПП"
        
        # Заголовки
        headers = ["№ п/п", "Уровень", "Код", "Наименование", "Статус", "Сумма", "Отделы"]
        ws.append(headers)
        
        # Стили заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Счетчик для нумерации
        counter = 0
        
        # Рекурсивный обход дерева с иерархической нумерацией
        def traverse(node: Dict, level: int = 0, prefix: str = ""):
            nonlocal counter
            counter += 1
            
            # Формируем иерархический номер
            if prefix:
                hierarchical_number = prefix
            else:
                hierarchical_number = str(counter)
            
            row = [
                counter,
                level,
                node.get("code", ""),
                f"{hierarchical_number} {node.get('name', '')}".strip(),
                node.get("status", ""),
                round(node.get("allocated", 0), 2),
                ", ".join(node.get("departments", []))
            ]
            ws.append(row)
            
            # Отступ для иерархии
            ws.cell(row=ws.max_row, column=4).alignment = Alignment(indent=level * 2)
            
            # Выделение корневых узлов жирным
            if level == 0:
                ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
            
            # Рекурсия для детей с иерархической нумерацией
            for idx, child in enumerate(node.get("children", []), 1):
                new_prefix = f"{hierarchical_number}.{idx}" if hierarchical_number else str(idx)
                traverse(child, level + 1, new_prefix)
        
        traverse(tree_data)
        
        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Границы для всех ячеек
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Сохранение
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
